"""Generate the standard FundLens Excel template for download."""

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from utils.data_loader import COLUMN_MAP

logger = logging.getLogger(__name__)

# Template columns derived from authoritative COLUMN_MAP (keeps ordering)
TEMPLATE_COLUMNS = list(COLUMN_MAP.keys())

# Example data row for guidance
EXAMPLE_ROW = [
    "2026-03-31", "支付宝", "主账户", "固收类", "纯债基金", "中国内地",
    "示例：XX纯债债券A", "000001", "CNY",
    12000, 11500, "", "", "",
    "中低", "稳健", 0.008, 0,
    "", "高", "（示例数据，可删除）",
]

HEADER_FILL = PatternFill(start_color="C96442", end_color="C96442", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FAF9F5")
EXAMPLE_FONT = Font(name="Arial", size=11, color="87867F")
BODY_FONT = Font(name="Arial", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="E8E6DC"),
    right=Side(style="thin", color="E8E6DC"),
    top=Side(style="thin", color="E8E6DC"),
    bottom=Side(style="thin", color="E8E6DC"),
)


def generate_template(output_path: str | Path) -> Path:
    """Create a standard FundLens asset snapshot template Excel file."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "asset_snapshot"

    # Header row
    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Example data row
    for col_idx, value in enumerate(EXAMPLE_ROW, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = EXAMPLE_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")

    # Set column widths
    col_widths = {
        1: 14, 2: 10, 3: 12, 4: 14, 5: 14, 6: 14,
        7: 24, 8: 16, 9: 8,
        10: 14, 11: 14, 12: 14, 13: 14, 14: 14,
        15: 10, 16: 10, 17: 16, 18: 16,
        19: 24, 20: 10, 21: 30,
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # Freeze header
    ws.freeze_panes = "A2"

    # Add instructions sheet
    ws2 = wb.create_sheet("使用说明")
    instructions = [
        ["FundLens 资产快照模板 — 填写说明", "", ""],
        ["", "", ""],
        ["字段", "是否必填", "说明"],
        ["统计日期", "必填", "格式 YYYY-MM-DD，如 2026-03-31"],
        ["平台", "必填", "支付宝 / 同花顺 / 其他"],
        ["账户名称", "选填", "如 主账户 / 定投账户"],
        ["资产大类", "建议填写", "现金类 / 固收类 / 固收增强类 / 权益类 / 跨境类 / 其他类"],
        ["产品类型", "建议填写", "如 纯债基金 / 固收+ / QDII / 宽基 ETF 等"],
        ["市场区域", "选填", "中国内地 / 香港 / 美国 / 全球 等"],
        ["产品名称", "必填", "基金/产品的完整名称"],
        ["产品代码", "选填", "6 位数字代码，系统自动补全前导零"],
        ["币种", "选填", "默认 CNY（人民币）"],
        ["当前金额", "必填", "唯一强制金额字段，填写当前市值"],
        ["持有成本", "选填", "不填则不参与收益率计算"],
        ["持有份额", "选填", "ETF 等有份额的产品可填写"],
        ["当前价格", "选填", "当前单位净值/价格"],
        ["成本价格", "选填", "买入时的单位净值/价格"],
        ["风险等级", "选填", "低 / 中低 / 中 / 中高 / 高"],
        ["资金用途", "选填", "活钱 / 稳健 / 长期 / 保障"],
        ["年化管理费率", "选填", "小数格式，如 0.008 表示 0.8%"],
        ["底层权益占比", "选填", "小数格式，如 0.8 表示 80%"],
        ["重仓行业（前3）", "选填", "用逗号分隔"],
        ["流动性", "选填", "高 / 中 / 低"],
        ["备注", "选填", "自由填写"],
    ]
    for row_idx, row_data in enumerate(instructions, start=1):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                cell.font = Font(name="Arial", size=16, bold=True, color="C96442")
            elif row_idx == 3:
                cell.font = Font(name="Arial", size=12, bold=True, color="141413")
                cell.fill = PatternFill(start_color="E8E6DC", end_color="E8E6DC", fill_type="solid")
            else:
                cell.font = Font(name="Arial", size=11, color="3D3D3A")

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 50

    wb.save(output_path)
    logger.info("Template generated at %s", output_path)
    return output_path


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    template_dir = Path(__file__).resolve().parent.parent / "data" / "sample"
    template_dir.mkdir(parents=True, exist_ok=True)
    path = generate_template(template_dir / "FundLens_Asset_Snapshot_Template.xlsx")
    print(f"Template created: {path}")
